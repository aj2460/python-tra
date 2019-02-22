import openpyxl as xl
wb=xl.load_workbook('test.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,2)
# print(cell.value)

# print(sheet.max_row)

for row in range (2, sheet.max_row+1):
    cell=sheet.cell(row, 2)
    print(cell.value)
    corrected_amount=cell.value*0.9
    corrected_amount_cell=sheet.cell(row,3)
    corrected_amount_cell.value=corrected_amount

wb.save('test2.xlsx')

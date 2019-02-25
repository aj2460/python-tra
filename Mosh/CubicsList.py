import openpyxl as xl
from  cursor import print_at
import ctypes

lib = ctypes.cdll.LoadLibrary("libwrapctype.dll")
lib.hello(' ')


class SalesMan():
    def __init__(self, name, comp, point):
        self.name = name
        self.comp = comp
        self.point = point


def col_to_list(col, rows):
    list = []
    for cell in col:
        c = cell[0]
        if c.value is None:
            break
        list.append(c.value)
    return list


wb = xl.load_workbook('name.xlsx')
# print(wb.sheetnames)
sheet = wb.active

sheet1 = wb['Sheet1']
number_of_rows = sheet1.max_row

name_col = sheet1['a1':'a' + str(number_of_rows)]
name_list = col_to_list(name_col, number_of_rows)

comp_col = sheet1['b1':'b' + str(number_of_rows)]
comp_list = col_to_list(comp_col, number_of_rows)

n_list=name_list[1:]
c_list=comp_list[1:]

n_list.sort()
c_list.sort()
for i in range(len(n_list)):
    print(f" ({i}) {n_list[i]}")
print("*" * 10)
for i in range(len(c_list)):
    print(f" ({i}) {c_list[i]}")

print (" ")
print ("-" *10)
print (" ")
emp = []
row = 2
col=20
for x in range(3):
    print_at(0,0," ")
    for i in range(len(n_list)):
        print(f" ({i}) {n_list[i]}")
    print("*" * 10)
    for i in range(len(c_list)):
        print(f" ({i}) {c_list[i]}")

    print(" ")
    print("-" * 10)
    print(" ")


    n = input('Enter the index of Salesman ')
    c = input('Enter the index of comp ')
    p = input('Enter the Points ')
    e = SalesMan(n_list[int(n)], c_list[int(c)], p)
    emp.append(e)
    lib.hello(' ')
    for e in emp:
        s = f"{e.name}  -  {e.comp}  : {e.point}"
        print_at(row, col, s)
        row += 1


# for e in emp:
#     print(f" {e.name}  -  {e.comp}  : {e.point}")

